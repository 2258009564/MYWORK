from openpyxl import load_workbook  #导入openpyxl模块 用于打开excel文件

my_workbook = load_workbook('python excel.xlsx')  #打开excel工作簿
sheet1 = my_workbook['Sheet1']  #获取工作表

sheet1.iter_rows(values_only=True)  #遍历工作表中的所有单元格
sheet1.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4)  #遍历指定单元格
'''
min_row=  最小行
max_row=  最大行
min_col=  最小列
max_col=  最大列
'''
